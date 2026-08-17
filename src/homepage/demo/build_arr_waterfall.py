#!/usr/bin/env python3
"""Build the ARR Waterfall Analysis workbook for the $100M -> $150M ARR demo.

Two blocks on one sheet, Pacer AI house style (white bg, navy/teal, Georgia-ish serif
headers rendered as Cambria which is the closest Excel-native serif to Cormorant):
  1. 3-Year ARR Waterfall (2023-2025) with a % of Beginning ARR column pair.
  2. Quarterly ARR Waterfall - 2025, with % of Beginning ARR, directly below.

Reads scenario.json so the numbers stay in lockstep with the chart + reel.
Run:  ./.venv/bin/python build_arr_waterfall.py
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SCEN = json.load(open(os.path.join(HERE, "scenario.json")))

# ---- Palette (Pacer AI house style) --------------------------------------
NAVY      = "FF0F1929"   # deep navy for section headers
NAVY_TXT  = "FF1C3560"
TEAL      = "FF27899A"
TEAL_LT   = "FF70C49C"
GREEN     = "FF2DB87A"
RED       = "FFC0504D"
GOLD      = "FFB8862B"
GREY_TXT  = "FF6B7280"
HAIR      = "FFD9DEE6"
BAND      = "FFF3F6F9"   # light row band
BEGIN_BG  = "FFEAF1F4"   # subtotal band (teal-tinted)
END_BG    = "FFDDEBEF"

SERIF = "Cambria"        # Excel-native serif standing in for Cormorant Garamond
SANS  = "Calibri"

thin  = Side(style="thin",  color=HAIR)
med   = Side(style="medium", color=NAVY_TXT)
top_border = Border(top=med)
box_hair   = Border(left=thin, right=thin, top=thin, bottom=thin)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "ARR Waterfall"
    ws.sheet_view.showGridLines = False

    # Column widths: A label, B..F periods, G spacer, H..L % columns
    ws.column_dimensions["A"].width = 30
    for c in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 13
    ws.column_dimensions["G"].width = 3
    for c in ["H", "I", "J", "K", "L"]:
        ws.column_dimensions[c].width = 11

    r = 1
    # ---- Title block -----------------------------------------------------
    ws.cell(r, 1, "Pacer AI").font = Font(name=SERIF, size=11, bold=True, color=TEAL)
    r += 1
    t = ws.cell(r, 1, "ARR Waterfall Analysis")
    t.font = Font(name=SERIF, size=22, bold=True, color=NAVY_TXT)
    r += 1
    s = ws.cell(r, 1, "Path from $100M to $150M ARR  ·  Illustrative $100M ARR SaaS  ·  $ in millions")
    s.font = Font(name=SANS, size=10, italic=True, color=GREY_TXT)
    r += 2

    annual = SCEN["annual"]
    a3 = [a for a in annual if a["year"] in ("2023", "2024", "2025")]

    r = block_annual(ws, r, "3-Year ARR Waterfall  ·  2023–2025", a3)
    r += 2
    r = block_quarterly(ws, r, "Quarterly ARR Waterfall  ·  2025", SCEN["quarterly_2025"])

    # ---- Footnote --------------------------------------------------------
    r += 2
    fn = ("Customer Base Growth = Expansion − Contraction − Gross Churn (= NRR − 100%).  "
          "New Logo ARR is net-new-customer ARR.  % columns are of Beginning ARR for the period.  "
          "2025 ending ARR ties to $100.0M; 2026–27 are illustrative guidance to $150M.")
    fc = ws.cell(r, 1, fn)
    fc.font = Font(name=SANS, size=8, italic=True, color=GREY_TXT)
    fc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=12)

    out = os.path.join(HERE, "ARR-Waterfall-Analysis.xlsx")
    wb.save(out)
    return out


def _money(ws, row, col, val, *, bold=False, color=None, sign=False, fill=None):
    c = ws.cell(row, col, round(val, 2))
    c.number_format = '#,##0.0;(#,##0.0)'
    c.font = Font(name=SANS, size=10.5, bold=bold, color=color or NAVY_TXT)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c


def _pct(ws, row, col, val, *, bold=False, color=None, fill=None):
    c = ws.cell(row, col, val / 100.0)
    c.number_format = '0.0%;(0.0%)'
    c.font = Font(name=SANS, size=10, bold=bold, color=color or GREY_TXT)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c


def _section_header(ws, row, text):
    for col in range(1, 13):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=NAVY)
    h = ws.cell(row, 1, text)
    h.font = Font(name=SERIF, size=13, bold=True, color="FFFFFFFF")
    h.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _col_headers(ws, row, periods):
    ws.cell(row, 1, "").font = Font(name=SANS, size=9)
    # $ header spanning
    dh = ws.cell(row, 2, "ARR movement ($M)")
    dh.font = Font(name=SANS, size=9, bold=True, color=TEAL)
    dh.alignment = Alignment(horizontal="left")
    ph = ws.cell(row, 8, "% of Beginning ARR")
    ph.font = Font(name=SANS, size=9, bold=True, color=TEAL)
    ph.alignment = Alignment(horizontal="left")
    row += 1
    hdr_font = Font(name=SANS, size=10, bold=True, color=NAVY_TXT)
    lab = ws.cell(row, 1, "")
    lab.font = hdr_font
    for i, p in enumerate(periods):
        c = ws.cell(row, 2 + i, p)
        c.font = hdr_font
        c.alignment = Alignment(horizontal="right")
        c.border = Border(bottom=Side(style="medium", color=TEAL))
        pc = ws.cell(row, 8 + i, p)
        pc.font = hdr_font
        pc.alignment = Alignment(horizontal="right")
        pc.border = Border(bottom=Side(style="medium", color=TEAL))
    return row + 1


def _row(ws, row, label, dollars, pcts, *, kind="mov"):
    """kind: 'begin' | 'mov' | 'sub' | 'end'"""
    band = None
    label_color = NAVY_TXT
    bold = False
    if kind == "begin":
        band = BEGIN_BG; bold = True
    elif kind == "end":
        band = END_BG; bold = True
    elif kind == "sub":
        bold = True
    lc = ws.cell(row, 1, label)
    lc.font = Font(name=SANS, size=10.5, bold=bold, color=label_color)
    lc.alignment = Alignment(indent=(0 if kind in ("begin", "end", "sub") else 1))
    if band:
        lc.fill = PatternFill("solid", fgColor=band)
        ws.cell(row, 7).fill = PatternFill("solid", fgColor="FFFFFFFF")
    for i, d in enumerate(dollars):
        color = NAVY_TXT
        if kind == "mov":
            color = GREEN if d >= 0 else RED
        _money(ws, row, 2 + i, d, bold=bold, color=color, fill=band)
    for i, p in enumerate(pcts):
        if p is None:
            continue
        color = GREY_TXT
        if kind == "mov":
            color = GREEN if p >= 0 else RED
        _pct(ws, row, 8 + i, p, bold=bold, color=color, fill=band)
    if kind == "end":
        for col in range(2, 13):
            cur = ws.cell(row, col).border
            ws.cell(row, col).border = Border(top=Side(style="medium", color=NAVY_TXT))
    return row + 1


def block_annual(ws, r, title, rows):
    _section_header(ws, r, title); r += 1
    periods = [a["year"] for a in rows]
    r = _col_headers(ws, r, periods)

    def series(key):
        return [a[key] for a in rows]

    begin = series("begin")
    new_logo = [a["begin"] * a["new_logo_pct"] / 100 for a in rows]
    expansion = [a["begin"] * a["expansion_pct"] / 100 for a in rows]
    contraction = [a["begin"] * a["contraction_pct"] / 100 for a in rows]
    churn = [a["begin"] * a["churn_pct"] / 100 for a in rows]
    base = [a["begin"] * a["base_pct"] / 100 for a in rows]
    end = series("end")

    r = _row(ws, r, "Beginning ARR", begin, [None] * len(rows), kind="begin")
    r = _row(ws, r, "New Logo ARR", new_logo, [a["new_logo_pct"] for a in rows], kind="mov")
    r = _row(ws, r, "Expansion ARR", expansion, [a["expansion_pct"] for a in rows], kind="mov")
    r = _row(ws, r, "Contraction ARR", contraction, [a["contraction_pct"] for a in rows], kind="mov")
    r = _row(ws, r, "Gross Churn ARR", churn, [a["churn_pct"] for a in rows], kind="mov")
    r = _row(ws, r, "  Customer Base Growth", base, [a["base_pct"] for a in rows], kind="sub")
    net_new = [end[i] - begin[i] for i in range(len(rows))]
    r = _row(ws, r, "Net New ARR", net_new, [a["total_pct"] for a in rows], kind="sub")
    r = _row(ws, r, "Ending ARR", end, [round(end[i] / begin[i] * 100, 1) for i in range(len(rows))], kind="end")

    # retention memo
    r += 1
    memo = ws.cell(r, 1, "Memo:  NRR / GRR")
    memo.font = Font(name=SANS, size=9, italic=True, color=GREY_TXT)
    for i, a in enumerate(rows):
        c = ws.cell(r, 2 + i, f'{a["nrr"]}% / {a["grr"]}%')
        c.font = Font(name=SANS, size=9, italic=True, color=TEAL)
        c.alignment = Alignment(horizontal="right")
    return r


def block_quarterly(ws, r, title, qs):
    _section_header(ws, r, title); r += 1
    periods = [q["period"] for q in qs]
    r = _col_headers(ws, r, periods)

    begin = [q["begin"] for q in qs]
    new_logo = [q["new_logo"] for q in qs]
    churn = [-q["begin"] * q["churn_rate"] / 100 for q in qs]
    contraction = [-q["begin"] * q["contraction_rate"] / 100 for q in qs]
    # expansion is the plug so base_net ties exactly
    base_net = [q["base_net"] for q in qs]
    expansion = [base_net[i] - churn[i] - contraction[i] for i in range(len(qs))]
    end = []
    for i, q in enumerate(qs):
        end.append(q["begin"] + new_logo[i] + base_net[i])

    def pct(vals):
        return [round(vals[i] / begin[i] * 100, 1) for i in range(len(qs))]

    r = _row(ws, r, "Beginning ARR", begin, [None] * len(qs), kind="begin")
    r = _row(ws, r, "New Logo ARR", new_logo, pct(new_logo), kind="mov")
    r = _row(ws, r, "Expansion ARR", expansion, pct(expansion), kind="mov")
    r = _row(ws, r, "Contraction ARR", contraction, pct(contraction), kind="mov")
    r = _row(ws, r, "Gross Churn ARR", churn, pct(churn), kind="mov")
    r = _row(ws, r, "  Customer Base Growth", base_net, pct(base_net), kind="sub")
    net_new = [end[i] - begin[i] for i in range(len(qs))]
    r = _row(ws, r, "Net New ARR", net_new, pct(net_new), kind="sub")
    r = _row(ws, r, "Ending ARR", end, pct(end), kind="end")
    return r


if __name__ == "__main__":
    path = build()
    # quick tie-out to stdout
    a = {x["year"]: x for x in SCEN["annual"]}
    print("Wrote:", path)
    print(f'  2025 ending ARR ties to ${a["2025"]["end"]:.1f}M  (target $100.0M)')
    print(f'  2027(f) ending ARR ${a["2027f"]["end"]:.2f}M  (target ~$150M)')
